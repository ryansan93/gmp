import glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DIR = os.path.dirname(os.path.abspath(__file__))
NUMF = '#,##0.00;(#,##0.00)'
ARIAL = 'Arial'
hdr_font = Font(name=ARIAL, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='1F4E78')
tot_font = Font(name=ARIAL, bold=True)
tot_fill = PatternFill('solid', fgColor='D9E1F2')
base_font = Font(name=ARIAL)
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(x):
    x = x.strip()
    return float(x) if x not in ('', 'NULL') else 0.0

units = ['BJN','BWI','GSK','JBR','KDR','LMG','LMJ','MDN','MGT','MJK','MLG','PRB','PSR','SLM','TAG']
cols = ['bym','tagihan','transfer','pemb','HR_credit','GL_main','GL_96010','GL_total','beda']

# load data
data = {}
for u in units:
    path = os.path.join(DIR, f'_{u}_byr_hr_vs_gl.txt')
    if not os.path.exists(path):
        path = os.path.join(DIR, f'_{u.lower()}_byr_hr_vs_gl.txt')
    rows = []
    with open(path, encoding='utf-8') as fh:
        for i, line in enumerate(fh):
            line = line.rstrip('\n')
            if not line or line.startswith('bym|'):
                continue
            p = line.split('|')
            if len(p) < 9:
                continue
            rows.append([p[0]] + [f(v) for v in p[1:9]])
    data[u] = rows

wb = Workbook()

# ---- Summary sheet ----
ws = wb.active
ws.title = 'Ringkasan'
title = ['Unit','Jml meleset','Pola1 (+) <1','Pola2 (-) <1','Net pembulatan','Item besar (n)','Item besar (Rp)','Net total']
ws.append(['RINGKASAN SELISIH HR vs GL — RHPP 21213 (per pembayaran BYM)'])
ws['A1'].font = Font(name=ARIAL, bold=True, size=13)
ws.append([])
ws.append(title)
for c in range(1, len(title)+1):
    cell = ws.cell(row=3, column=c); cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = border

r = 4
for u in units:
    rows = data[u]
    n = len(rows)
    pola1 = sum(x[8] for x in rows if 0 < x[8] < 1)
    pola2 = sum(x[8] for x in rows if -1 < x[8] < 0)
    big = [x for x in rows if abs(x[8]) >= 1]
    netp = pola1 + pola2
    nett = sum(x[8] for x in rows)
    ws.append([u, n, round(pola1,2), round(pola2,2), round(netp,2), len(big), round(sum(x[8] for x in big),2), round(nett,2)])
    r += 1
# total row
last = r-1
def col_total(idx):
    return round(sum(data[u_] and 0 for u_ in []) , 2)  # placeholder
ws.append(['TOTAL',
           sum(len(data[u]) for u in units),
           round(sum(x[8] for u in units for x in data[u] if 0 < x[8] < 1),2),
           round(sum(x[8] for u in units for x in data[u] if -1 < x[8] < 0),2),
           round(sum(x[8] for u in units for x in data[u] if abs(x[8])<1),2),
           sum(1 for u in units for x in data[u] if abs(x[8])>=1),
           round(sum(x[8] for u in units for x in data[u] if abs(x[8])>=1),2),
           round(sum(x[8] for u in units for x in data[u]),2)])
for c in range(1, len(title)+1):
    cell = ws.cell(row=r, column=c); cell.font = tot_font; cell.fill = tot_fill; cell.border = border
for rr in range(4, r+1):
    for c in range(1, len(title)+1):
        cell = ws.cell(row=rr, column=c)
        if rr != r:
            cell.font = base_font
        cell.border = border
        if c >= 3 and c != 6:
            cell.number_format = NUMF
ws.append([])
note = ('Catatan: net (utk 12 unit tanpa item besar) = kolom SELISIH laporan live PERSIS. '
        'Item besar = kasus BUKAN pembulatan yg SUDAH ditangani cap/kompensasi di laporan live '
        '(MLG BYM/03/26/00081 M.Ali dobel; SLM BYM/03/26/00263 THORIQ; LMG overpay).')
ws.cell(row=r+2, column=1, value=note).font = Font(name=ARIAL, italic=True, size=9)
ws.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=8)
ws.cell(row=r+2, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[r+2].height = 45
widths = [10,12,14,14,16,14,16,16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = 'A4'

# ---- per-unit sheets ----
for u in units:
    s = wb.create_sheet(u)
    s.append(cols)
    for c in range(1, len(cols)+1):
        cell = s.cell(row=1, column=c); cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    for row in data[u]:
        s.append(row)
    n = len(data[u])
    # total row (computed values)
    tr = n + 2
    s.cell(row=tr, column=1, value='TOTAL')
    for ci in range(2, len(cols)+1):
        s.cell(row=tr, column=ci, value=round(sum(row[ci-1] for row in data[u]), 2))
    for c in range(1, len(cols)+1):
        s.cell(row=tr, column=c).font = tot_font
        s.cell(row=tr, column=c).fill = tot_fill
    # formatting
    for rr in range(2, tr+1):
        for c in range(1, len(cols)+1):
            cell = s.cell(row=rr, column=c)
            if rr != tr:
                cell.font = base_font
            cell.border = border
            if c >= 2:
                cell.number_format = NUMF
    s.column_dimensions['A'].width = 18
    for col in 'BCDEFGHI':
        s.column_dimensions[col].width = 15
    s.freeze_panes = 'A2'

out = os.path.join(DIR, 'selisih_rhpp.xlsx')
wb.save(out)
print('saved', out)
